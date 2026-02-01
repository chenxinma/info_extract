import re
import json
import logging
from pathlib import Path
from typing import Any, AsyncGenerator, Literal, TypedDict, TypeAlias
import warnings
from openpyxl import load_workbook
from openpyxl.styles.proxy import StyleProxy
import pandas as pd

from ..pipeline import Step, StepResult
from ..config.profile_manager import ProfileManager
from ..utils import classify_excel_sheets, excel_to_png_via_com

logger = logging.getLogger(__name__)
warnings.simplefilter("ignore", category=UserWarning)

SheetType: TypeAlias = Literal["Data", "Form", "Unknown"]


class SheetMeta(TypedDict):
    sheet_name: str
    header_row: int
    columns: list[str]
    sample_data: list[dict[str, Any]]


class ExcelReader(Step):
    """处理Excel文件，提取sheet内容"""

    def __init__(self, source_dir: str = "source", processing_dir: str = "processing"):
        """
        初始化Excel处理器

        Args:
            source_dir: 源文件目录路径
            processing_dir: 处理后文件保存目录路径
        """
        self.source_dir = Path(source_dir)
        self.processing_dir = Path(processing_dir)

        # 确保目录存在
        self.source_dir.mkdir(exist_ok=True)
        self.processing_dir.mkdir(exist_ok=True)
        self.meta_list: list[SheetMeta] = []

    async def run(
        self, profile_manager: ProfileManager
    ) -> AsyncGenerator[StepResult, None]:
        """
        处理Excel文件xls或xlsx，提取sheet内容

        Yields:
            按sheet返回parquest文件路径
        """
        _files = self.source_files(self.source_dir, "*.xls*")
        logger.info(f"找到 {len(_files)} 个excel文件")

        for excel_file in _files:
            if excel_file.stem.endswith("~"):
                continue

            # 分类sheet
            sheets = classify_excel_sheets(str(excel_file))
            if not sheets:
                logger.warning(f"分类sheet失败 {excel_file}")
                continue

            # 处理sheet
            logger.info(f"处理 {excel_file} 中的 {len(sheets)} 个sheet")
            data_sheets = []
            form_sheets = []
            for sheet in sheets:
                sheet_type: SheetType = sheet["sheet_type"]
                if sheet_type == "Data":
                    data_sheets.append(sheet["original"]["sheet_name"])
                elif sheet_type == "Form":
                    form_sheets.append(sheet["original"]["sheet_name"])

            try:
                book = load_workbook(str(excel_file))
            except Exception as exp:
                logger.warning(f"加载excel文件 {excel_file} 异常 {str(exp)}")
                continue
            for sheet_name in data_sheets:
                sheet = book[sheet_name]
                # 检查sheet是否为隐藏状态，如果是则跳过
                if sheet.sheet_state == "hidden" or sheet.sheet_state == "veryHidden":
                    logger.debug(f"sheet {sheet_name} 是隐藏状态，已跳过处理")
                    continue

                # header_row = self.find_header_row(sheet, header_candidates=["姓名","身份证"])
                header_row = self.improved_find_header_row(
                    sheet, header_candidates=None
                )
                if header_row == -1:
                    continue

                fname = f"{excel_file.stem}_{sheet_name}"
                # 读取数据
                df = pd.read_excel(
                    excel_file, sheet_name=sheet_name, header=header_row, dtype=str
                )
                if (
                    "姓名" in df.columns
                ):  # TODO: 改进，在配置中标明关键列，代替此处hardcode 的姓名
                    df.dropna(subset=["姓名"], inplace=True)  # 姓名为空的行，删除

                n_rows = len(df)
                if n_rows == 0:
                    logger.warning(f"sheet {fname} 无数据")
                    continue
                # 添加行颜色列
                df["行颜色"] = self.fetch_row_colors(sheet, header_row, n_rows)
                meta = SheetMeta(
                    sheet_name=fname,
                    header_row=header_row,
                    columns=list(df.columns),
                    sample_data=df.sample(min(3, n_rows)).to_dict(orient="records"),  # type: ignore
                )
                self.meta_list.append(meta)
                parquet_file = self.processing_dir / f"{fname}.parquet"
                df.to_parquet(parquet_file, index=False)
                yield str(parquet_file), excel_file.stem

        with open(self.processing_dir / "excel_meta", "w", encoding="utf-8") as f:
            json.dump(self.meta_list, f, ensure_ascii=False, indent=4)


    def fetch_row_colors(self, sheet, header_row: int, row_count: int) -> list[str]:
        """
        获取指定行的背景颜色

        Args:
            sheet: 工作表对象
            header_row: 表头行号（从0开始）
            row_count: 要获取的行数

        Returns:
            行的背景颜色（RGB值），如果没有颜色则返回"无颜色"
        """
        # 获取数据行的背景色
        row_colors = []
        # 注意：dataframe的索引从0开始，而Excel行号从1开始，且需要跳过表头行
        for row_idx in range(
            header_row + 2, min(header_row + 2 + row_count, sheet.max_row + 1)
        ):
            # 获取整行第一个非空单元格的填充色（通常整行颜色一致）
            fill_color = None
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None and str(cell.value).strip():
                    if isinstance(cell.fill, StyleProxy) and cell.fill.fgColor.rgb:
                        if (
                            cell.fill.fgColor.indexed == 64
                            or cell.fill.fgColor.rgb == "00000000"
                            or str(cell.fill.fgColor.rgb)
                            == "Values must be of type <class 'str'>"
                        ):
                            continue

                        fill_color = str(cell.fill.fgColor.rgb)[2:]
                    break
            row_colors.append(fill_color if fill_color else "")

        # 确保row_colors长度与row_count一致
        if len(row_colors) < row_count:
            row_colors.extend([""] * (row_count - len(row_colors)))
        elif len(row_colors) > row_count:
            row_colors = row_colors[:row_count]

        return row_colors

    def find_header_row(
        self, sheet, header_candidates=None, max_scan_rows=10, blanks_ratio=0.5
    ):
        """
        在 Excel 工作表中自动定位“表头”所在行号。

        Args
        ----
        sheet : xlrd.sheet.Sheet
            已经打开的工作表
        header_candidates : list[str] | None
            如果你已经知道表头里必须出现哪些关键字，可以写进来，
            例如 ["姓名","年龄"]，函数会优先返回同时包含这些关键字的第一行。
            留 None 则按“空值比例”策略自动判断。
        max_scan_rows : int
            最多从前 N 行里扫描，避免整张表都扫一遍
        blanks_ratio : float
            允许的行空值比例阈值，小于该比例才认为可能是表头
            （0.5 表示“该行一半以上有值”就合格）

        Returns
        -------
        int
            表头行号（从 0 开始），找不到则返回 -1
        """
        ncols = max(sheet.max_column, 1)
        for r_idx, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=min(max_scan_rows, sheet.max_row)),
            start=0,
        ):
            row_values = [cell.value for cell in row]
            # 1. 空值比例策略
            non_blanks = sum(1 for v in row_values if v is not None and str(v).strip())
            if non_blanks / ncols < (1 - blanks_ratio):
                continue

            # 2. 关键字策略（如果调用者给了关键字列表）
            if header_candidates:
                hit = all(
                    any(key in str(cv).strip() for cv in row_values)
                    for key in header_candidates
                )
                if not hit:
                    continue
            # 同时满足两个策略，就认为找到了
            return r_idx
        return -1

    def improved_find_header_row(
        self,
        sheet,
        header_candidates=None,
        max_scan_rows=10,
        blanks_ratio=0.5,
        fuzzy_match_threshold=0.8,
    ):
        """
        改进的表头行查找算法
        新增特征：文本属性（中文/名词）、数字占比、与数据行的区分度、唯一值密度
        """
        # 预处理：获取所有扫描行的原始值和处理后的值（去空、去空格）
        scan_rows = []
        max_row = min(max_scan_rows, sheet.max_row)
        ncols = max(sheet.max_column, 1) if sheet.max_column else 1  # 处理空表格
        max_row_value_count = 0
        for r_idx in range(1, max_row + 1):
            row_cells = sheet[r_idx]  # 按行索引取单元格（从1开始）
            row_values = [cell.value for cell in row_cells]
            # 处理后的值：非空、去空格、转字符串
            processed_vals = [
                str(v).strip()
                for v in row_values
                if v is not None and str(v).strip() != ""
            ]
            scan_rows.append(
                {
                    "r_idx": r_idx - 1,  # 原始返回的是0开始的索引
                    "raw_values": row_values,
                    "processed": processed_vals,
                    "non_blank_count": len(processed_vals),
                }
            )
            if len(row_values) > max_row_value_count:
                max_row_value_count = len(row_values)

        # 不对单列数据进行处理
        if max_row_value_count < 2:
            return -1

        # 遍历每一行，计算综合评分
        header_scores = []
        for row_data in scan_rows:
            r_idx = row_data["r_idx"]
            raw_vals = row_data["raw_values"]
            processed_vals = row_data["processed"]
            non_blank_count = row_data["non_blank_count"]
            score = 0.0

            # 规则1：空值比例过滤（不满足则直接跳过，评分为0）
            blank_ratio = 1 - (non_blank_count / ncols) if ncols > 0 else 1.0
            if blank_ratio > blanks_ratio:
                header_scores.append((r_idx, 0.0))
                continue
            score += 1.0  # 满足空值比例，基础分

            # 规则2：唯一值比例（表头通常唯一值多）
            unique_vals = set(processed_vals)
            unique_ratio = (
                len(unique_vals) / len(processed_vals) if processed_vals else 0.0
            )
            score += unique_ratio * 1.0  # 权重1.0

            # 规则3：数字占比（表头通常数字少，数据行数字多）
            num_count = 0
            for val in processed_vals:
                # 判断是否为纯数字（整数/小数，排除身份证等长数字字符串）
                if (
                    re.match(r"^-?\d+(\.\d+)?$", val) and len(val) <= 15
                ):  # 15位以内纯数字视为数值
                    num_count += 1
            num_ratio = num_count / len(processed_vals) if processed_vals else 1.0
            score += (1 - num_ratio) * 1.5  # 数字占比越低，得分越高，权重1.5

            # 规则4：中文字符占比（表头通常包含中文，数据行可能少）
            chinese_count = 0
            for val in processed_vals:
                chinese_count += sum(1 for c in val if "\u4e00" <= c <= "\u9fff")
            chinese_ratio = (
                chinese_count / sum(len(val) for val in processed_vals)
                if processed_vals
                else 0.0
            )
            score += chinese_ratio * 2.0  # 权重2.0（中文对表头识别更重要）

            # 规则5：下方行的数据验证（表头下方应该是数据行，满足数据特征）
            # 取当前行下一行（如果存在），判断是否为数据行（数字占比高、非空）
            next_row_idx = r_idx + 1
            if next_row_idx < len(scan_rows):
                next_row = scan_rows[next_row_idx]
                next_processed = next_row["processed"]
                next_non_blank = next_row["non_blank_count"]
                if next_non_blank / ncols >= (1 - blanks_ratio):  # 下一行非空比例足够
                    # 计算下一行的数字占比
                    next_num_count = 0
                    for val in next_processed:
                        if re.match(r"^-?\d+(\.\d+)?$", val) and len(val) <= 15:
                            next_num_count += 1
                    next_num_ratio = (
                        next_num_count / len(next_processed) if next_processed else 0.0
                    )
                    if next_num_ratio > 0.5:  # 下一行数字占比超过50%，视为数据行
                        score += 1.5  # 权重1.5

            # 规则6：关键字模糊匹配（如果有候选关键字）
            if header_candidates:
                matched_count = 0
                for candidate in header_candidates:
                    candidate_lower = candidate.lower().strip()
                    # 模糊匹配：候选词在单元格值中（忽略大小写）
                    if any(
                        candidate_lower in str(v).lower().strip()
                        for v in raw_vals
                        if v is not None
                    ):
                        matched_count += 1
                match_ratio = (
                    matched_count / len(header_candidates) if header_candidates else 0.0
                )
                if match_ratio >= fuzzy_match_threshold:
                    score += match_ratio * 3.0  # 权重3.0（关键字匹配优先级最高）
                else:
                    score *= 0.5  # 匹配不足，降低评分

            header_scores.append((r_idx, score))

        # 找到评分最高的行（如果多个最高分，取第一个）
        if not header_scores:
            return -1
        header_scores.sort(key=lambda x: (-x[1], x[0]))  # 按评分降序、索引升序
        best_r_idx, best_score = header_scores[0]

        # 如果最高分是0，说明没有符合条件的行
        return best_r_idx if best_score > 0 else -1
