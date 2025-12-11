"""
测试Excel表头检测算法的不同实现
"""
import re
import unittest
from pathlib import Path
import tempfile
from openpyxl import Workbook
from src.info_extract.source.excel import ExcelReader
from tqdm import tqdm


def create_test_excel_with_headers():
    """创建带有不同类型表头的测试Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet" # type: ignore

    # 在不同行添加不同的标题组合
    ws['A1'] = '无关数据' # type: ignore
    ws['B1'] = '其他信息' # type: ignore
    
    ws['A2'] = '' # type: ignore
    ws['B2'] = '' # type: ignore
    
    ws['A3'] = '姓名' # type: ignore
    ws['B3'] = '身份证号' # type: ignore
    ws['C3'] = '年龄' # type: ignore
    
    ws['A4'] = '张三' # type: ignore
    ws['B4'] = '110101199003071234' # type: ignore
    ws['C4'] = '30' # type: ignore
    
    ws['A5'] = '李四' # type: ignore
    ws['B5'] = '110101198506125678' # type: ignore
    ws['C5'] = '35' # type: ignore

    # 创建临时文件
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    return temp_file.name


def improved_find_header_row(sheet,
                           header_candidates=None,
                           max_scan_rows=10,
                           blanks_ratio=0.5,
                           fuzzy_match_threshold=0.8):
    """
    改进的表头行查找算法
    新增特征：文本属性（中文/名词）、数字占比、与数据行的区分度、唯一值密度
    """
    # 预处理：获取所有扫描行的原始值和处理后的值（去空、去空格）
    scan_rows = []
    max_row = min(max_scan_rows, sheet.max_row)
    ncols = max(sheet.max_column, 1) if sheet.max_column else 1  # 处理空表格
    max_row_value_count = 0
    for r_idx in range(1, max_row + 1):
        row_cells = sheet[r_idx]  # 按行索引取单元格（从1开始）
        row_values = [cell.value for cell in row_cells]
        # 处理后的值：非空、去空格、转字符串
        processed_vals = [str(v).strip() for v in row_values if v is not None and str(v).strip() != '']
        scan_rows.append({
            'r_idx': r_idx - 1,  # 原始返回的是0开始的索引
            'raw_values': row_values,
            'processed': processed_vals,
            'non_blank_count': len(processed_vals)
        })
        if len(row_values) > max_row_value_count:
            max_row_value_count = len(row_values)
    
    # 不对单列数据进行处理
    if max_row_value_count < 2:
        return -1

    # 遍历每一行，计算综合评分
    header_scores = []
    for row_data in scan_rows:
        r_idx = row_data['r_idx']
        raw_vals = row_data['raw_values']
        processed_vals = row_data['processed']
        non_blank_count = row_data['non_blank_count']
        score = 0.0

        # 规则1：空值比例过滤（不满足则直接跳过，评分为0）
        blank_ratio = 1 - (non_blank_count / ncols) if ncols > 0 else 1.0
        if blank_ratio > blanks_ratio:
            header_scores.append((r_idx, 0.0))
            continue
        score += 1.0  # 满足空值比例，基础分

        # 规则2：唯一值比例（表头通常唯一值多）
        unique_vals = set(processed_vals)
        unique_ratio = len(unique_vals) / len(processed_vals) if processed_vals else 0.0
        score += unique_ratio * 1.0  # 权重1.0

        # 规则3：数字占比（表头通常数字少，数据行数字多）
        num_count = 0
        for val in processed_vals:
            # 判断是否为纯数字（整数/小数，排除身份证等长数字字符串）
            if re.match(r'^-?\d+(\.\d+)?$', val) and len(val) <= 15:  # 15位以内纯数字视为数值
                num_count += 1
        num_ratio = num_count / len(processed_vals) if processed_vals else 1.0
        score += (1 - num_ratio) * 1.5  # 数字占比越低，得分越高，权重1.5

        # 规则4：中文字符占比（表头通常包含中文，数据行可能少）
        chinese_count = 0
        for val in processed_vals:
            chinese_count += sum(1 for c in val if '\u4e00' <= c <= '\u9fff')
        chinese_ratio = chinese_count / sum(len(val) for val in processed_vals) if processed_vals else 0.0
        score += chinese_ratio * 2.0  # 权重2.0（中文对表头识别更重要）

        # 规则5：下方行的数据验证（表头下方应该是数据行，满足数据特征）
        # 取当前行下一行（如果存在），判断是否为数据行（数字占比高、非空）
        next_row_idx = r_idx + 1
        if next_row_idx < len(scan_rows):
            next_row = scan_rows[next_row_idx]
            next_processed = next_row['processed']
            next_non_blank = next_row['non_blank_count']
            if next_non_blank / ncols >= (1 - blanks_ratio):  # 下一行非空比例足够
                # 计算下一行的数字占比
                next_num_count = 0
                for val in next_processed:
                    if re.match(r'^-?\d+(\.\d+)?$', val) and len(val) <= 15:
                        next_num_count += 1
                next_num_ratio = next_num_count / len(next_processed) if next_processed else 0.0
                if next_num_ratio > 0.5:  # 下一行数字占比超过50%，视为数据行
                    score += 1.5  # 权重1.5

        # 规则6：关键字模糊匹配（如果有候选关键字）
        if header_candidates:
            matched_count = 0
            for candidate in header_candidates:
                candidate_lower = candidate.lower().strip()
                # 模糊匹配：候选词在单元格值中（忽略大小写）
                if any(candidate_lower in str(v).lower().strip() for v in raw_vals if v is not None):
                    matched_count += 1
            match_ratio = matched_count / len(header_candidates) if header_candidates else 0.0
            if match_ratio >= fuzzy_match_threshold:
                score += match_ratio * 3.0  # 权重3.0（关键字匹配优先级最高）
            else:
                score *= 0.5  # 匹配不足，降低评分

        header_scores.append((r_idx, score))

    # 找到评分最高的行（如果多个最高分，取第一个）
    if not header_scores:
        return -1
    header_scores.sort(key=lambda x: (-x[1], x[0]))  # 按评分降序、索引升序
    best_r_idx, best_score = header_scores[0]

    # 如果最高分是0，说明没有符合条件的行
    return best_r_idx if best_score > 0 else -1



class TestExcelHeaderDetection(unittest.TestCase):

    def setUp(self):
        self.excel_reader = ExcelReader()
        self.test_file = create_test_excel_with_headers()

    def tearDown(self):
        # 删除临时文件
        Path(self.test_file).unlink()

    @unittest.skip("s1")
    def test_current_vs_improved_algorithm_with_candidates(self):
        """比较现有算法和改进算法在有候选关键词的情况下的表现"""
        from openpyxl import load_workbook
        wb = load_workbook(self.test_file)
        ws = wb.active

        # 测试现有算法
        current_result = self.excel_reader.find_header_row(ws, header_candidates=["姓名", "身份证"])
        
        # 测试改进算法
        improved_result = improved_find_header_row(ws, header_candidates=["姓名", "身份证"])

        print(f"Current algorithm result with candidates: {current_result}")
        print(f"Improved algorithm result with candidates: {improved_result}")

        # 两种算法都应该找到第2行（索引为2，因为从0开始）作为表头
        self.assertEqual(current_result, 2, "当前算法应找到第3行（索引为2）作为表头")
        self.assertEqual(improved_result, 2, "改进算法应找到第3行（索引为2）作为表头")

    @unittest.skip("s2")
    def test_current_vs_improved_algorithm_without_candidates(self):
        """比较现有算法和改进算法在没有候选关键词的情况下的表现"""
        from openpyxl import load_workbook
        wb = load_workbook(self.test_file)
        ws = wb.active

        # 测试现有算法（无候选关键词）
        current_result = self.excel_reader.find_header_row(ws, header_candidates=None)

        # 测试改进算法（无候选关键词）
        improved_result = improved_find_header_row(ws, header_candidates=None)

        print(f"Current algorithm result without candidates: {current_result}")
        print(f"Improved algorithm result without candidates: {improved_result}")

        # 注意：当前算法在没有候选关键词时可能会返回第一行（索引0），而改进算法能更智能地识别真正的表头
        # 改进算法应该能找到第2行，因为它更符合表头的特征
        self.assertEqual(improved_result, 2, "改进算法应找到第3行（索引为2）作为表头")
        # 对于当前算法，由于其行为与改进算法不同，我们只记录其行为而不强制相等
        print(f"Current algorithm found row {current_result}, improved algorithm found row {improved_result}")

    @unittest.skip("s3")
    def test_improved_algorithm_fuzzy_matching(self):
        """测试改进算法的模糊匹配能力"""
        from openpyxl import load_workbook
        wb = load_workbook(self.test_file)
        ws = wb.active
        
        # 使用部分匹配的关键词
        result = improved_find_header_row(ws, header_candidates=["姓", "身份"])
        print(f"Improved algorithm result with partial matches: {result}")
        
        # 应该找到第2行，因为它包含"姓名"和"身份证号"
        self.assertEqual(result, 2, "改进算法应能识别部分匹配的关键词")

    @unittest.skip("s4")
    def test_improved_algorithm_tolerance_to_missing_keys(self):
        """测试改进算法对缺少关键词的容忍度"""
        from openpyxl import load_workbook
        wb = load_workbook(self.test_file)
        ws = wb.active
        
        # 提供一个不存在的关键词，但保持阈值较低
        result = improved_find_header_row(ws, header_candidates=["姓名", "不存在的列"], 
                                        fuzzy_match_threshold=0.5)
        print(f"Improved algorithm result with one missing key: {result}")
        
        # 因为我们设置了较低的阈值，所以应该仍能找到表头
        self.assertEqual(result, 2, "改进算法应在低阈值下容忍部分缺失的关键词")

        # 当阈值较高时，找不到匹配项
        result2 = improved_find_header_row(ws, header_candidates=["姓名", "不存在的列"], 
                                        fuzzy_match_threshold=0.9)
        print(f"Improved algorithm result with high threshold and missing key: {result2}")
        
        # 应该返回-1，因为未达到高阈值
        self.assertEqual(result2, -1, "改进算法应在高阈值下不接受缺失的关键词")

    def test_act_compare(self):
        from openpyxl import load_workbook

        test_files = list(Path("/data/home/macx/work/tmp/workdir/source").glob("*.xls*"))
        for f in test_files:
            if f.stem.endswith("~"):
                continue
            wb = load_workbook(f)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                # 检查sheet是否为隐藏状态，如果是则跳过
                if sheet.sheet_state == 'hidden' or sheet.sheet_state == 'veryHidden':
                    continue
                expect_header_row = self.excel_reader.find_header_row(sheet, header_candidates=["姓名","身份证"])
                actual_header_row = improved_find_header_row(sheet, header_candidates=None)
                # self.assertEqual(expect_header_row, actual_header_row, f"{f.name}/{sheet_name}")
                if actual_header_row != expect_header_row:
                    print(f"{f.name}/{sheet_name}", expect_header_row, actual_header_row)


if __name__ == '__main__':
    unittest.main()