from typing import override
import unittest
from unittest.case import skip

from openpyxl import load_workbook
from info_extract.source import ExcelReader
import pandas as pd

class TestExcelReader(unittest.TestCase):
    @override
    def setUp(self) -> None:
        self.excel_reader = ExcelReader(source_dir=r"D:\tmp\info_extract\work\source", 
                                        processing_dir=r"D:\tmp\info_extract\work\processing")
        return super().setUp()
        
    def test_fetch_row_colors(self):
        book = load_workbook(r"D:\tmp\info_extract\work\source\达美乐离职申报表-青岛.xlsx")
        sheet = book["2025-10"]

        header_row = self.excel_reader.find_header_row(sheet=sheet, header_candidates=['姓名', '身份证'])
        colors = self.excel_reader.fetch_row_colors(sheet=sheet, header_row=header_row, row_count=4)
        print(colors)
    
    def test_fetch_row_colors2(self):
        book = load_workbook(r"D:\tmp\info_extract\work\source\社保减员附件.xlsx")
        sheet = book["0"]

        header_row = self.excel_reader.find_header_row(sheet=sheet, header_candidates=['姓名', '身份证'])
        colors = self.excel_reader.fetch_row_colors(sheet=sheet, header_row=header_row, row_count=7)
        print(colors)
    
    def test_get_columns(self):
        excel_file = r"D:\tmp\info_extract\work\source\达美乐离职申报表-青岛.xlsx"
        sheet_name = "2025-10"
        book = load_workbook(excel_file)
        sheet = book[sheet_name]

        header_row = self.excel_reader.find_header_row(sheet=sheet, header_candidates=['姓名', '身份证'])
        df = pd.read_excel(excel_file, 
                           sheet_name=sheet_name, 
                           header=header_row,
                           dtype=str)
        print(df.columns)